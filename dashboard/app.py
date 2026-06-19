import streamlit as st
import os
import io
import csv
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Telecom KPI Automation Dashboard", layout="wide")
st.title("📡 Telecom KPI Automation Dashboard")

# ----------------------------------------
# KPIs shown in Pivot sheet
# ----------------------------------------
PIVOT_KPIS = [
    "4G_Total VoLte Traffic_24 Hrs",
    "4G_Total Payload (Data)_24 Hrs",
]

CELL_COL_CANDIDATES = [
    "Nokia_4G.NAME", "Cell Name", "Cell", "CellName", "Short name",
    "Site", "NE Name", "Object", "LNBTS name", "eNB Name", "Cell ID",
    "NE", "Network Element", "eNodeB Name", "eNB", "Node",
]

SKIP_CELL_VALUES = {
    "none", "nan", "", "active", "inactive", "total",
    "grand total", "grand", "summary", "subtotal", "all",
}

PIVOT_SHEET = "KPI Pivot"

# ----------------------------------------
# Session state — master workbook lives in memory
# ----------------------------------------
if "wb_bytes" not in st.session_state:
    st.session_state.wb_bytes = None   # raw bytes of the master xlsx


def load_wb_from_state():
    """Load workbook from session state bytes, or create a fresh one."""
    if st.session_state.wb_bytes:
        return load_workbook(io.BytesIO(st.session_state.wb_bytes))
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def save_wb_to_state(wb):
    """Save workbook back to session state as bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    st.session_state.wb_bytes = buf.getvalue()
    return buf.getvalue()


# ----------------------------------------
# Helpers
# ----------------------------------------
def smart_cast(v):
    v = str(v).strip()
    try:    return int(v)
    except: pass
    try:    return float(v)
    except: pass
    return v


def extract_date_from_filename(filename, default_month="June", default_year=2026):
    import re

    # 1) Numeric date formats: 2026-06-01, 20260601, 01-06-2026, 01_06_2026
    for pat, fmt in [
        (r"(\d{4}-\d{2}-\d{2})", "%Y-%m-%d"),
        (r"(\d{8})",              "%Y%m%d"),
        (r"(\d{2}-\d{2}-\d{4})", "%d-%m-%Y"),
        (r"(\d{2}_\d{2}_\d{4})", "%d_%m_%Y"),
    ]:
        m = re.search(pat, filename)
        if m:
            try:    return datetime.strptime(m.group(1), fmt).strftime("%d %B,%Y")
            except: pass

    # 2) Ordinal-word formats: "1st june", "2nd June", "23rd_june", "june11", "june 11"
    name = filename.lower()

    # Strip file extension
    name = re.sub(r"\.(csv|xlsx|xls)$", "", name)

    # Pattern: <day><suffix> <month>   e.g. "1st june", "23rd_june"
    m = re.search(r"(\d{1,2})(st|nd|rd|th)?[\s_\-]*([a-z]+)", name)
    if m:
        day_str, _, month_str = m.groups()
        try:
            day = int(day_str)
            # Try to match month name (partial match, e.g. "june" -> June)
            for month_num in range(1, 13):
                month_name = datetime(2000, month_num, 1).strftime("%B").lower()
                if month_str.startswith(month_name[:3]):  # match first 3 letters
                    dt = datetime(default_year, month_num, day)
                    return dt.strftime("%d %B,%Y")
        except (ValueError, ):
            pass

    # Pattern: <month><day>  e.g. "june11", "june 16"
    m2 = re.search(r"([a-z]+)[\s_\-]*(\d{1,2})", name)
    if m2:
        month_str, day_str = m2.groups()
        try:
            day = int(day_str)
            for month_num in range(1, 13):
                month_name = datetime(2000, month_num, 1).strftime("%B").lower()
                if month_str.startswith(month_name[:3]):
                    dt = datetime(default_year, month_num, day)
                    return dt.strftime("%d %B,%Y")
        except (ValueError, ):
            pass

    # 3) Nothing matched — fall back to today, but warn the user
    st.warning(
        f"⚠️ Could not parse a date from filename '{filename}'. "
        f"Using today's date as fallback — please verify this is correct!"
    )
    return datetime.today().strftime("%d %B,%Y")


def parse_csv(reader):
    """Find real header row; return (raw_header, cell_col_name, cell_idx, data_rows)."""
    for ri, row in enumerate(reader):
        vals = [str(c).strip() for c in row]
        for name in CELL_COL_CANDIDATES:
            if name in vals:
                return vals, name, vals.index(name), reader[ri + 1:]
    return None, None, None, []


def build_norm_rows(raw_header, cell_idx, data_rows, master_header, date_label):
    wanted = []
    for i, col in enumerate(raw_header):
        if i == cell_idx: continue
        if not col: continue
        if col.startswith("Day:"): continue
        wanted.append((col, i))

    csv_lookup = {col: i for col, i in wanted}

    if master_header is None:
        master_header = ["Date", "Cell Name"] + [c for c, _ in wanted]
    else:
        existing = set(master_header)
        for col, i in wanted:
            if col not in existing:
                master_header.append(col)

    master_idx   = {col: i for i, col in enumerate(master_header)}
    hdr_lower    = {c.lower() for c in master_header} | {c.lower() for c in CELL_COL_CANDIDATES}

    norm_rows = []
    for row in data_rows:
        if not any(str(c).strip() for c in row): continue
        cell_val = str(row[cell_idx]).strip() if cell_idx < len(row) else ""
        if not cell_val: continue
        if cell_val.lower() in SKIP_CELL_VALUES: continue
        if cell_val.lower() in hdr_lower: continue
        try:    float(cell_val); continue
        except: pass

        out = [""] * len(master_header)
        out[0] = date_label
        out[1] = cell_val
        for col, mi in master_idx.items():
            if col in ("Date", "Cell Name"): continue
            ci = csv_lookup.get(col)
            if ci is not None and ci < len(row):
                out[mi] = smart_cast(row[ci])
        norm_rows.append(out)

    return master_header, norm_rows


def rebuild_pivot(wb):
    pivot_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    pivot_seen = defaultdict(lambda: defaultdict(lambda: defaultdict(bool)))
    all_dates, all_cells = set(), set()

    for v in ["Nokia", "Ericsson", "Samsung", "Huawei"]:
        if v not in wb.sheetnames: continue
        ws = wb[v]

        all_rows = [[str(c).strip() if c is not None else "" for c in row]
                    for row in ws.iter_rows(values_only=True)]
        if len(all_rows) < 2: continue

        header = None; header_pos = 0
        for ri, row in enumerate(all_rows):
            if "Date" in row and "Cell Name" in row:
                header = row; header_pos = ri; break
        if header is None: continue

        date_idx = header.index("Date")
        cell_idx = header.index("Cell Name")
        kpi_col  = {kpi: header.index(kpi) for kpi in PIVOT_KPIS if kpi in header}
        if not kpi_col: continue

        for row in all_rows[header_pos + 1:]:
            if not any(row): continue
            if row[0].lower() == "date": continue
            date_val = row[date_idx] if date_idx < len(row) else ""
            cell_val = row[cell_idx] if cell_idx < len(row) else ""
            if not date_val or not cell_val: continue
            if cell_val.lower() in SKIP_CELL_VALUES: continue
            all_dates.add(date_val); all_cells.add(cell_val)
            for kpi, ci in kpi_col.items():
                if ci < len(row) and row[ci] not in ("", "None"):
                    try:
                        pivot_data[cell_val][kpi][date_val] += float(row[ci])
                        pivot_seen[cell_val][kpi][date_val]  = True
                    except: pass

    sorted_dates = sorted(all_dates)
    sorted_cells = sorted(all_cells)

    if PIVOT_SHEET in wb.sheetnames: del wb[PIVOT_SHEET]
    wp = wb.create_sheet(title=PIVOT_SHEET)

    kpi_fill   = PatternFill("solid", start_color="1F4E79")
    date_fill  = PatternFill("solid", start_color="2E75B6")
    total_fill = PatternFill("solid", start_color="BDD7EE")
    label_fill = PatternFill("solid", start_color="D6E4F0")
    alt_fill   = PatternFill("solid", start_color="EBF3FB")
    white_fill = PatternFill("solid", start_color="FFFFFF")
    bw     = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    bd     = Font(bold=True, color="1F4E79", name="Calibri", size=10)
    blabel = Font(bold=True, name="Calibri", size=10)
    norm   = Font(name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    n = len(sorted_dates)
    def cs(ki): return 2 + ki * (n + 1)

    wp.cell(row=1, column=1, value="").fill = kpi_fill
    for ki, kpi in enumerate(PIVOT_KPIS):
        c1 = cs(ki)
        if n > 1:
            wp.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1+n-1)
        c = wp.cell(row=1, column=c1, value=f"Sum of {kpi}")
        c.fill = kpi_fill; c.font = bw; c.alignment = center
        tc = wp.cell(row=1, column=c1+n, value=f"Total Sum of {kpi}")
        tc.fill = kpi_fill; tc.font = bw; tc.alignment = center

    rl = wp.cell(row=2, column=1, value="Cell Name")
    rl.fill = date_fill; rl.font = bw; rl.alignment = center
    for ki in range(len(PIVOT_KPIS)):
        c1 = cs(ki)
        for di, date in enumerate(sorted_dates):
            c = wp.cell(row=2, column=c1+di, value=date)
            c.fill = date_fill; c.font = bw; c.alignment = center
        gc = wp.cell(row=2, column=c1+n, value="(blank)")
        gc.fill = total_fill; gc.font = bd; gc.alignment = center

    wp.freeze_panes = "B3"

    for ri, cell_name in enumerate(sorted_cells):
        er = ri + 3
        fill = alt_fill if ri % 2 == 0 else white_fill
        lc = wp.cell(row=er, column=1, value=cell_name)
        lc.fill = label_fill; lc.font = blabel; lc.alignment = left; lc.border = bdr
        for ki, kpi in enumerate(PIVOT_KPIS):
            c1 = cs(ki); total = 0.0; has = False
            for di, date in enumerate(sorted_dates):
                c = wp.cell(row=er, column=c1+di)
                if pivot_seen[cell_name][kpi].get(date):
                    v = round(pivot_data[cell_name][kpi][date], 6)
                    c.value = v; total += v; has = True
                else:
                    c.value = ""
                c.font = norm; c.fill = fill; c.alignment = center; c.border = bdr
            gc = wp.cell(row=er, column=c1+n)
            gc.value = round(total, 6) if has else ""
            gc.font = blabel; gc.fill = total_fill; gc.alignment = center; gc.border = bdr

    wp.column_dimensions["A"].width = 26
    for ci in range(2, 2 + len(PIVOT_KPIS) * (n+1)):
        wp.column_dimensions[get_column_letter(ci)].width = 16
    wp.row_dimensions[1].height = 45
    wp.row_dimensions[2].height = 30


# ============================================================
# UI
# ============================================================

st.sidebar.header("📁 Master File")

# Upload existing master file to continue previous session
uploaded_master = st.sidebar.file_uploader(
    "Load existing master file (optional)", type=["xlsx"], key="master_upload"
)
if uploaded_master is not None:
    st.session_state.wb_bytes = uploaded_master.read()
    st.sidebar.success("✅ Master file loaded!")

# Download master file
if st.session_state.wb_bytes:
    st.sidebar.download_button(
        label="⬇️ Download Master Excel",
        data=st.session_state.wb_bytes,
        file_name="Telecom_Master.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.sidebar.info("No master file yet. Upload a CSV to create one.")

st.divider()

# ----------------------------------------
# Vendor + CSV upload
# ----------------------------------------
vendor = st.selectbox("Select Vendor", ["Nokia", "Ericsson", "Samsung", "Huawei"])

input_method = st.radio(
    "How do you want to provide the CSV?",
    ["Upload CSV file", "Upload ZIP (recommended for large files)", "Local file path (run app locally)"],
    horizontal=True
)

uploaded_file = None
uploaded_zip  = None
local_path    = None

if input_method == "Upload CSV file":
    uploaded_file = st.file_uploader("Upload Daily KPI CSV", type=["csv"])
elif input_method == "Upload ZIP (recommended for large files)":
    st.caption("Zip your CSV first — typically 70-80% smaller, much faster & more reliable upload.")
    uploaded_zip = st.file_uploader("Upload Zipped CSV", type=["zip"])
else:
    local_path = st.text_input(
        "Paste full file path",
        placeholder="C:/Users/irina/OneDrive/Documents/NOKIA_KPI_AUTOMATION/input/file.csv"
    )
    if local_path and not os.path.exists(local_path):
        st.error("File not found. Check the path.")
        local_path = None

file_ready = (uploaded_file is not None or uploaded_zip is not None
              or (local_path and os.path.exists(local_path)))

if file_ready:
    if st.button("Update Master Dataset"):
        try:
            import zipfile

            # Determine filename + get a text stream ready for csv.reader
            if uploaded_zip is not None:
                uploaded_zip.seek(0)
                zf = zipfile.ZipFile(io.BytesIO(uploaded_zip.read()))
                csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                if not csv_names:
                    st.error("No CSV file found inside the zip.")
                    st.stop()
                inner_name  = csv_names[0]
                fname       = inner_name
                raw_bytes   = zf.read(inner_name)
                text_stream = io.TextIOWrapper(io.BytesIO(raw_bytes), encoding="utf-8", errors="ignore")
            elif uploaded_file is not None:
                fname = uploaded_file.name
                uploaded_file.seek(0)
                text_stream = io.TextIOWrapper(uploaded_file, encoding="utf-8", errors="ignore")
            else:
                fname       = os.path.basename(local_path)
                text_stream = open(local_path, encoding="utf-8", errors="ignore")

            date_label = extract_date_from_filename(fname)
            wb         = load_wb_from_state()

            if vendor in wb.sheetnames:
                ws = wb[vendor]
            else:
                ws = wb.create_sheet(title=vendor)

            existing_header = None
            if ws.max_row >= 1:
                r1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in r1):
                    existing_header = [str(v).strip() if v is not None else "" for v in r1]

            is_first = existing_header is None

            csv_iter = csv.reader(text_stream)

            raw_header    = None
            cell_idx      = None
            master_header = None
            csv_lookup    = {}
            master_idx    = {}
            hdr_lower     = {}
            header_found  = False
            written       = 0

            for row in csv_iter:
                vals = [str(c).strip() for c in row]

                # Step 1: find the header row
                if not header_found:
                    for name in CELL_COL_CANDIDATES:
                        if name in vals:
                            raw_header   = vals
                            cell_idx     = vals.index(name)
                            header_found = True
                            break
                    if not header_found:
                        continue

                    # Build master header immediately from the header row
                    wanted = [
                        (col, i) for i, col in enumerate(raw_header)
                        if i != cell_idx and col and not col.startswith("Day:")
                    ]
                    csv_lookup = {col: i for col, i in wanted}

                    if existing_header is None:
                        master_header = ["Date", "Cell Name"] + [c for c, _ in wanted]
                    else:
                        master_header = list(existing_header)
                        existing_set  = set(master_header)
                        for col, _ in wanted:
                            if col not in existing_set:
                                master_header.append(col)

                    master_idx = {col: i for i, col in enumerate(master_header)}
                    hdr_lower  = ({c.lower() for c in master_header} |
                                  {c.lower() for c in CELL_COL_CANDIDATES})

                    if is_first:
                        ws.append(master_header)
                    elif len(master_header) > len(existing_header or []):
                        for ci, val in enumerate(master_header, start=1):
                            ws.cell(row=1, column=ci, value=val)
                    continue   # header row itself is not a data row

                # Step 2: process data rows
                if not any(vals): continue
                cell_val = vals[cell_idx] if cell_idx < len(vals) else ""
                if not cell_val: continue
                if cell_val.lower() in SKIP_CELL_VALUES: continue
                if cell_val.lower() in hdr_lower: continue
                try:    float(cell_val); continue
                except: pass

                out = [""] * len(master_header)
                out[0] = date_label
                out[1] = cell_val
                for col, mi in master_idx.items():
                    if col in ("Date", "Cell Name"): continue
                    ci2 = csv_lookup.get(col)
                    if ci2 is not None and ci2 < len(vals):
                        out[mi] = smart_cast(vals[ci2])
                ws.append(out)
                written += 1

            if raw_header is None:
                st.error(f"Could not find cell column. Tried: {CELL_COL_CANDIDATES}")
                st.stop()

            if is_first:
                st.info(f"📝 Header written: {len(master_header)} columns")
            elif master_header and existing_header and len(master_header) > len(existing_header):
                st.info(f"📝 Header extended to {len(master_header)} columns")

            # Delete stray header rows (rows 3+)
            to_del = [
                i for i in range(3, ws.max_row + 1)
                if ws.cell(i, 1).value is not None
                and str(ws.cell(i, 1).value).strip().lower() == "date"
            ]
            for i in reversed(to_del):
                if i > 1: ws.delete_rows(i)

            # Close local file if used
            if local_path:
                text_stream.close()

            rebuild_pivot(wb)
            xlsx_bytes = save_wb_to_state(wb)

            st.success(f"✅ {vendor}: {written} rows | {len(master_header)} cols | {date_label}")
            if to_del:
                st.info(f"🧹 {len(to_del)} stray header rows removed")

            st.download_button(
                label="⬇️ Download Updated Master Excel",
                data=xlsx_bytes,
                file_name="Telecom_Master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            import traceback
            st.error(str(e)); st.code(traceback.format_exc())

# ----------------------------------------
# Delete a date's data
# ----------------------------------------
st.divider()
if st.session_state.wb_bytes:
    with st.expander("🗑️ Delete incorrect data"):
        wb_del = load_wb_from_state()
        available_sheets = [s for s in ["Nokia", "Ericsson", "Samsung", "Huawei"]
                            if s in wb_del.sheetnames]
        if available_sheets:
            del_vendor = st.selectbox("Select vendor sheet", available_sheets, key="del_vendor")
            ws_del = wb_del[del_vendor]

            # Collect all unique dates from that sheet
            all_rows_del = [[str(c).strip() if c is not None else ""
                             for c in row] for row in ws_del.iter_rows(values_only=True)]
            # Find header
            date_col_idx = 0
            for row in all_rows_del:
                if "Date" in row:
                    date_col_idx = row.index("Date")
                    break
            unique_dates = sorted(set(
                row[date_col_idx] for row in all_rows_del[1:]
                if row and row[date_col_idx] and row[date_col_idx].lower() != "date"
            ))

            if unique_dates:
                del_date = st.selectbox("Select date to delete", unique_dates, key="del_date")
                if st.button("🗑️ Delete this date's rows", type="primary"):
                    # Delete all rows where col[date_col_idx] == del_date (reverse order)
                    to_del = [
                        i for i in range(2, ws_del.max_row + 1)
                        if str(ws_del.cell(i, date_col_idx + 1).value).strip() == del_date
                    ]
                    for i in reversed(to_del):
                        ws_del.delete_rows(i)
                    rebuild_pivot(wb_del)
                    xlsx_bytes = save_wb_to_state(wb_del)
                    st.success(f"✅ Deleted {len(to_del)} rows for {del_date} from {del_vendor}")
                    st.download_button(
                        label="⬇️ Download Updated Master Excel",
                        data=xlsx_bytes,
                        file_name="Telecom_Master.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_delete"
                    )
            else:
                st.info("No dates found in this sheet.")
        else:
            st.info("No vendor sheets found. Upload a CSV first.")

# ----------------------------------------
# Refresh Pivot
# ----------------------------------------
st.divider()
if st.session_state.wb_bytes:
    if st.button("🔄 Refresh KPI Pivot Sheet"):
        try:
            wb = load_wb_from_state()
            rebuild_pivot(wb)
            xlsx_bytes = save_wb_to_state(wb)
            st.success("✅ KPI Pivot refreshed!")
            st.download_button(
                label="⬇️ Download Updated Master Excel",
                data=xlsx_bytes,
                file_name="Telecom_Master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_refresh"
            )
        except Exception as e:
            st.error(str(e))

# ----------------------------------------
# Preview Vendor Sheet
# ----------------------------------------
if st.session_state.wb_bytes:
    wb = load_wb_from_state()
    if vendor in wb.sheetnames:
        ws = wb[vendor]
        st.divider()
        st.subheader(f"{vendor} Sheet Preview")
        all_rows = list(ws.values)
        header_row = None; header_pos = 0
        for ri, row in enumerate(all_rows):
            row_s = [str(c).strip() if c is not None else "" for c in row]
            if "Date" in row_s and "Cell Name" in row_s:
                header_row = row_s; header_pos = ri; break
        if header_row:
            display = [header_row] + [
                [str(c) if c is not None else "" for c in r]
                for r in all_rows[header_pos + 1:]
                if any(c is not None for c in r)
            ]
            import pandas as pd
            
            # De-duplicate column names to support PyArrow engine formatting requirements
            seen_cols = {}
            unique_vendor_headers = []
            for col_name in display[0]:
                if col_name in seen_cols:
                    seen_cols[col_name] += 1
                    unique_vendor_headers.append(f"{col_name}_{seen_cols[col_name]}")
                else:
                    seen_cols[col_name] = 0
                    unique_vendor_headers.append(col_name)

            df = pd.DataFrame(display[1:], columns=unique_vendor_headers)
            df = df.astype(str).replace("None", "")
            st.dataframe(df, use_container_width=True)
        else:
            import pandas as pd
            df = pd.DataFrame([[str(c) if c is not None else "" for c in r] for r in all_rows])
            st.dataframe(df, use_container_width=True)
        col1, col2 = st.columns(2)
        col1.metric("Rows", ws.max_row)
        col2.metric("Columns", ws.max_column)
    else:
        st.info("No data uploaded for this vendor yet.")

# ----------------------------------------
# Preview KPI Pivot Sheet
# ----------------------------------------
if st.session_state.wb_bytes:
    wb = load_wb_from_state()
    if PIVOT_SHEET in wb.sheetnames:
        st.divider()
        st.subheader("📊 KPI Pivot Preview")
        import pandas as pd
        pivot_rows = list(wb[PIVOT_SHEET].values)
        if pivot_rows:
            # De-duplicate Pivot column names safely
            seen_pivot_cols = {}
            unique_pivot_headers = []
            for c in pivot_rows[0]:
                c_str = str(c) if c is not None else ""
                if c_str in seen_pivot_cols:
                    seen_pivot_cols[c_str] += 1
                    unique_pivot_headers.append(f"{c_str}_{seen_pivot_cols[c_str]}")
                else:
                    seen_pivot_cols[c_str] = 0
                    unique_pivot_headers.append(c_str)

            pdf = pd.DataFrame(
                [[str(c) if c is not None else "" for c in r] for r in pivot_rows[1:]],
                columns=unique_pivot_headers
            )
            st.dataframe(pdf, use_container_width=True)